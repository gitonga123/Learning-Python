# -*- coding: utf-8 -*-

#download and install the openpyxl library. Linux PIP INSTALL openpyxl

#import the library in your code
import openpyxl

#open the work sheet, store the the work sheet in the same folder as your file.py
# you can change the path by os.chdir()

wb = openpyxl.load_workbook('texting.xlsx')

#print the name of the work sheets you have
print wb.get_sheet_names()

#print the name of the work sheet you want
sheet = wb.get_sheet_by_name('Sheet1')

#print the title
type(sheet)
print sheet.title

#select active worksheet
anotherSheet = wb.active

print anotherSheet


#print the content of a cell
print sheet['A1']
print sheet['A1'].value

c = sheet['B1'].value

print c

#select a specific cell to work with

d = sheet.cell(row=2, column=2).value
print d

#print the content automatical
for i in range(1,8):
	print i,sheet.cell(row=i, column=1).value, "\t",  sheet.cell(row=i, column=2).value

#print the size of the work sheet

print ""
print sheet.max_row

print sheet.max_column